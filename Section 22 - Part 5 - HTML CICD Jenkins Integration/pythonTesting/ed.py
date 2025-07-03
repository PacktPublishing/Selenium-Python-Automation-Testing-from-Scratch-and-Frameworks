import openpyxl
from selenium.webdriver.chrome import webdriver
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


# book = openpyxl.load_workbook(file_path)
# sheet = book.active
# Dict = {}

def read_excel_data(file_path, searchTerm, colName, new_value):
    book = openpyxl.load_workbook(file_path)
    sheet = book.active
    Dict = {}

    # Find "Apple" and related data
    for i in range(1, sheet.max_row + 1):
        for j in range(1, sheet.max_column + 1):
            if sheet.cell(row=i, column=j).value == searchTerm:
                Dict["row"] = i
                break

    for i in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=i).value == colName:
            Dict["col"] = i

    sheet.cell(row=Dict["row"], column=Dict["col"]).value = new_value
    book.save(file_path)


file_path = "/Users/rahulshetty/downloads/download.xlsx"
driver = webdriver.Chrome()

driver.implicitly_wait(5)  # Set implicit wait for 5 seconds

driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")

# Click the download button
driver.find_element(By.ID, "downloadButton").click()
read_excel_data(file_path, "Apple", "price", "947")
# Upload the file
file_input = driver.find_element(By.CSS_SELECTOR, "input[type=file]")
file_input.send_keys(file_path)  # Replace with your file path

# Wait for toast visibility
toast_locator = (By.CSS_SELECTOR, ".Toastify__toast-body div:nth-child(2)")
WebDriverWait(driver, 10).until(EC.visibility_of_element_located(toast_locator))

# Optional: Get and verify toast text
toast_text = driver.find_element(*toast_locator).text
print("Toast text:", toast_text)
assert toast_text == "Updated Excel Data Successfully."

# Wait for toast to disappear
WebDriverWait(driver, 10).until(EC.invisibility_of_element_located(toast_locator))

# Get column ID
col_id = driver.find_element(By.XPATH, "//div[text()='Price']").get_attribute("data-column-id")

# Get actual price
actual_price = driver.find_element(By.XPATH,
                                   f"//div[text()='Apple']/parent::div/parent::div/div[@id='cell-{col_id}-undefined']").text
print("Actual price:", actual_price)

assert actual_price == "947"

driver.quit()  # Close the browser
